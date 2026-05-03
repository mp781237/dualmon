"""
動能資產配置自動更新腳本 (usdn_updater.py)
─────────────────────────────────────────
使用 yfinance 抓取 ETF 月度收盤價（日線 resample 到月底），計算五大策略訊號，並輸出：
  1. signals.json — 供 動能資產配置.html 儀表板讀取
  2. usdn.xlsx    — 若存在則同步股價分頁（可選）

使用方式：
    pip install yfinance openpyxl
    python usdn_updater.py

股價來源：yfinance adjusted close（含股息再投資與分割調整，近似 total return）
資料窗口：上月底前 N 個完成月（不含當月 partial bar，避免 1m 報酬被噪音污染）
"""

import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

import yfinance as yf

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# ── 設定 ──────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "usdn.xlsx"
JSON_PATH = SCRIPT_DIR / "signals.json"
ETF_LIST = ["VOO", "VXUS", "BND", "BIL", "VSS", "TLT", "VWO", "LQD", "IEF", "SHY", "QQQ"]
MONTHS_NEEDED = 13


def fetch_monthly_closes(ticker: str, months: int = MONTHS_NEEDED) -> list[tuple]:
    # end 設為本月 1 號（exclusive）→ 永遠只取「已收盤完成的月」
    # 避開 yfinance interval=1mo 在月初跑時回傳 partial 當月 bar 造成 1m 報酬失真
    end = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start = end - timedelta(days=(months + 2) * 35)  # 多抓 2 個月當緩衝

    tk = yf.Ticker(ticker)
    # 用日線抓 + auto_adjust=True 取調整後 close（含股息與分割），再 resample 到月底
    # 比 interval="1mo" 穩定（不受 yfinance 月線是否含 partial 行為差異影響）
    # 不開 repair=True：那需要 scipy（+~100MB）；對主流 ETF 資料品質夠用
    hist = tk.history(
        start=start.strftime("%Y-%m-%d"),
        end=end.strftime("%Y-%m-%d"),
        interval="1d",
        auto_adjust=True,
        actions=True,
    )

    if hist.empty:
        raise RuntimeError(f"{ticker}: yfinance returned empty data")

    if hist.index.tz is not None:
        hist.index = hist.index.tz_localize(None)

    # resample 到月底（"ME" = Month End），取每月最後一個交易日的 close
    closes = hist["Close"].resample("ME").last().dropna().tail(months)

    results = [(d.to_pydatetime().replace(day=1), float(c))
               for d, c in closes.items()]
    results.sort(key=lambda x: x[0], reverse=True)
    return results


def update_excel(wb, etf: str, data: list[tuple]):
    # Excel 欄位佈局（保留給下游 OHLC 公式參考，目前 OHLC 全填 close）：
    #   col 1 = 月份日期（每月 1 號）
    #   col 2 = 月底 close
    #   col 3-5 = open/high/low（沒抓 OHLC，全填 close 當 placeholder）
    #   col 6-7 = 保留欄位
    if etf not in wb.sheetnames:
        print(f"  [略過] 工作表 '{etf}' 不存在")
        return

    ws = wb[etf]

    for i, (date_val, close_val) in enumerate(data):
        r = i + 2
        ws.cell(row=r, column=1, value=date_val)
        ws.cell(row=r, column=2, value=close_val)
        for c in (3, 4, 5):
            ws.cell(row=r, column=c, value=close_val)
        for c in (6, 7):
            ws.cell(row=r, column=c, value="")

    for r in range(len(data) + 2, ws.max_row + 1):
        for c in range(1, 8):
            ws.cell(row=r, column=c, value=None)


def calc_returns(prices: list[float]) -> dict:
    if len(prices) < 2:
        return {}

    def ret(current, past):
        if past and past != 0:
            return (current - past) / past
        return None

    p = prices
    result = {}
    result["1m"] = ret(p[0], p[1]) if len(p) > 1 else None
    result["3m"] = ret(p[0], p[3]) if len(p) > 3 else None
    result["6m"] = ret(p[0], p[6]) if len(p) > 6 else None
    result["12m"] = ret(p[0], p[12]) if len(p) > 12 else None

    if all(result.get(k) is not None for k in ["1m", "3m", "6m", "12m"]):
        # VAA 13612W 公式（Keller & Keuning 2017）
        result["vaa"] = 12 * result["1m"] + 4 * result["3m"] + 2 * result["6m"] + result["12m"]
    if all(result.get(k) is not None for k in ["1m", "3m", "6m"]):
        # 注意：這不是真正的「加速度」(d momentum / dt)，是 1m/3m/6m 動能的簡單平均
        # 變數名 accel 為歷史命名，前端 subtitle 已誠實標示「分數 = avg(1m, 3m, 6m)」
        result["accel"] = (result["1m"] + result["3m"] + result["6m"]) / 3

    return result


def _classify_pick(pick: str) -> tuple[str, str]:
    """Return (mode_class, default_label) for a ticker."""
    if pick in ("VOO", "QQQ"):
        return ("attack", "美國股市")
    if pick in ("VXUS", "VSS", "VWO"):
        return ("intl", "外國股市")
    if pick in ("BND", "TLT", "LQD", "IEF"):
        return ("bond", "債券")
    if pick in ("BIL", "SHY"):
        return ("defense", "現金")
    return ("attack", "")


def _round(v, d=6):
    return round(v, d) if v is not None else None


def compute_signals(all_returns: dict) -> dict:
    def r(etf, key):
        return all_returns.get(etf, {}).get(key)

    def pct(v):
        return f"{v*100:+.2f}%" if v is not None else ""

    strategies = []

    # S1 經典雙動能
    voo_12, vxus_12, bil_12 = r("VOO", "12m"), r("VXUS", "12m"), r("BIL", "12m")
    if None not in (voo_12, vxus_12, bil_12):
        winner = "VOO" if voo_12 >= vxus_12 else "VXUS"
        pick = winner if max(voo_12, vxus_12) > bil_12 else "BND"
        mode, default_label = _classify_pick(pick)
        label = "債券避險" if pick == "BND" else default_label
        strategies.append({
            "id": "S1", "name": "原版雙動能",
            "pick": pick, "assets": ["VOO", "VXUS", "BND", "BIL"],
            "mode": mode, "modeLabel": label,
        })

    # S2 拉風增強
    qqq_12 = r("QQQ", "12m")
    if None not in (qqq_12, vxus_12, bil_12):
        winner = "QQQ" if qqq_12 >= vxus_12 else "VXUS"
        pick = winner if max(qqq_12, vxus_12) > bil_12 else "BND"
        mode, default_label = _classify_pick(pick)
        label = "債券避險" if pick == "BND" else default_label
        strategies.append({
            "id": "S2", "name": "原版雙動能 · 拉風增強版",
            "pick": pick, "assets": ["QQQ", "VXUS", "BND", "BIL"],
            "mode": mode, "modeLabel": label,
        })

    # S3 加速雙動能
    voo_acc, vss_acc, tlt_1m = r("VOO", "accel"), r("VSS", "accel"), r("TLT", "1m")
    if voo_acc is not None and vss_acc is not None:
        # max() 取最大值，平手時取插入順序的第一個（VOO）
        candidates = {"VOO": voo_acc, "VSS": vss_acc}
        winner, top_score = max(candidates.items(), key=lambda kv: kv[1])
        if top_score > 0:
            pick, score = winner, top_score
        elif tlt_1m is not None and tlt_1m < 0:
            pick, score = "BIL", None
        else:
            pick, score = "TLT", None
        mode, _ = _classify_pick(pick)
        if pick in ("VOO", "VSS"):
            label = f"進攻 {pct(score)}"
        elif pick == "TLT":
            label = "債券避險"
        else:
            label = "現金"
        strategies.append({
            "id": "S3", "name": "加速雙動能",
            "pick": pick, "assets": ["VOO", "VSS", "TLT", "BIL"],
            "mode": mode, "modeLabel": label,
        })

    # S4 騷速雙動能
    qqq_acc = r("QQQ", "accel")
    if qqq_acc is not None and vss_acc is not None:
        candidates = {"QQQ": qqq_acc, "VSS": vss_acc}
        winner, top_score = max(candidates.items(), key=lambda kv: kv[1])
        if top_score > 0:
            pick, score = winner, top_score
        elif tlt_1m is not None and tlt_1m < 0:
            pick, score = "BIL", None
        else:
            pick, score = "TLT", None
        mode, _ = _classify_pick(pick)
        if pick in ("QQQ", "VSS"):
            label = f"進攻 {pct(score)}"
        elif pick == "TLT":
            label = "債券避險"
        else:
            label = "現金"
        strategies.append({
            "id": "S4", "name": "騷速雙動能 · 含 QQQ",
            "pick": pick, "assets": ["QQQ", "VSS", "TLT", "BIL"],
            "mode": mode, "modeLabel": label,
        })

    # S5 VAA 攻擊型（客製化版本：defensive universe 加 BIL，標準 Keller 2017 為 LQD/IEF/SHY 3 檔）
    atk = {"VOO": r("VOO", "vaa"), "VXUS": r("VXUS", "vaa"),
           "VWO": r("VWO", "vaa"), "BND": r("BND", "vaa")}
    defs = {"LQD": r("LQD", "vaa"), "IEF": r("IEF", "vaa"),
            "SHY": r("SHY", "vaa"), "BIL": r("BIL", "vaa")}
    vaa_block = None
    if all(v is not None for v in atk.values()):
        all_pos = all(v > 0 for v in atk.values())
        if all_pos:
            pick = max(atk, key=atk.get)
            pick_label = "攻擊模式 · 最強標的"
            pick_reason = "ALL 攻擊資產得分 > 0"
            mode_label = "攻擊模式"
        else:
            valid_defs = {k: v for k, v in defs.items() if v is not None}
            if not valid_defs:
                raise RuntimeError("VAA defense universe missing data — cannot pick safely")
            pick = max(valid_defs, key=valid_defs.get)
            pick_label = "防禦模式 · 最強標的"
            pick_reason = "攻擊資產有負值，切防禦"
            mode_label = "防禦模式"
        mode, _ = _classify_pick(pick)
        strategies.append({
            "id": "S5", "name": "VAA 攻擊型",
            "pick": pick, "assets": list(atk.keys()),
            "mode": mode, "modeLabel": mode_label,
        })
        vaa_block = {
            "attack": [
                {"ticker": k, "score": _round(v, 4), "winner": all_pos and k == pick}
                for k, v in atk.items()
            ],
            "defense": [
                {"ticker": k, "score": _round(v, 4), "winner": (not all_pos) and k == pick}
                for k, v in defs.items()
            ],
            "pickLabel": pick_label,
            "pickReason": pick_reason,
            "pick": pick,
        }

    # 動能分數表（S3/S4 共用）
    accel_tickers = ["VOO", "QQQ", "VSS", "TLT", "BIL"]
    attack_accel = {t: r(t, "accel") for t in ("VOO", "QQQ", "VSS")}
    valid_attack = {k: v for k, v in attack_accel.items() if v is not None and v > 0}
    winner_ticker = max(valid_attack, key=valid_attack.get) if valid_attack else None
    momentum_rows = [{
        "ticker": t,
        "m1": _round(r(t, "1m")),
        "m3": _round(r(t, "3m")),
        "m6": _round(r(t, "6m")),
        "m12": _round(r(t, "12m")),
        "score": _round(r(t, "accel")),
        "winner": t == winner_ticker,
    } for t in accel_tickers]

    return {
        "updated": datetime.now().strftime("%Y-%m-%d"),
        "strategies": strategies,
        "momentum": {
            "title": "加速雙動能 · 騷速雙動能",
            "subtitle": "分數 = avg(1m, 3m, 6m)",
            "rows": momentum_rows,
        },
        "vaa": vaa_block,
    }


def write_signals_json(data: dict) -> None:
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已儲存訊號至 {JSON_PATH}")


def print_signal(data: dict):
    def fmt_pct(v):
        return f"{v*100:+.2f}%" if v is not None else "N/A"

    print("\n" + "=" * 60)
    print("  動能資產配置 — 當月策略訊號")
    print("=" * 60)

    for s in data["strategies"]:
        print(f"\n  {s['id']} {s['name']}:  {s['pick']}  ({s['modeLabel']})")

    print("\n  動能分數（加速分數）:")
    for row in data["momentum"]["rows"]:
        print(f"     {row['ticker']:4s}  "
              f"1m={fmt_pct(row['m1']):>9s}  "
              f"3m={fmt_pct(row['m3']):>9s}  "
              f"6m={fmt_pct(row['m6']):>9s}  "
              f"accel={fmt_pct(row['score']):>9s}"
              f"{'  ★' if row['winner'] else ''}")

    if data["vaa"]:
        print("\n  VAA 攻擊資產分數:")
        for item in data["vaa"]["attack"]:
            mark = " ★" if item["winner"] else ""
            print(f"     {item['ticker']:4s}  {item['score']:+.4f}{mark}")

    print("\n" + "=" * 60)


def main():
    print("動能資產配置更新程式 v2.1")
    print(f"JSON 輸出: {JSON_PATH}")
    print(f"XLSX 目標: {XLSX_PATH} {'(存在)' if XLSX_PATH.exists() else '(不存在 — 略過)'}")
    print(f"更新時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("-" * 40)

    all_data = {}
    all_returns = {}

    failed = []
    for etf in ETF_LIST:
        print(f"  抓取 {etf}...", end=" ", flush=True)
        try:
            data = fetch_monthly_closes(etf)
        except Exception as e:
            print(f"失敗 ({e})")
            failed.append(etf)
            continue
        all_data[etf] = data
        prices = [d[1] for d in data]
        all_returns[etf] = calc_returns(prices)
        print(f"OK ({len(data)} 個月)")

    # M4: fail-fast — 任何 ETF 抓不到就 abort，避免靜默產出 partial signals.json
    if failed:
        print(f"\n[錯誤] 以下 ETF 抓取失敗：{failed}")
        sys.exit(1)

    # 計算一次 signals 給 JSON 與 console 共用（O1：避免重複呼叫 compute_signals）
    print()
    data = compute_signals(all_returns)

    # 1. 寫 JSON（主要輸出）
    write_signals_json(data)

    # 2. 同步 Excel（可選）
    if XLSX_PATH.exists() and load_workbook is not None:
        print("\n更新 Excel 檔案...")
        wb = load_workbook(str(XLSX_PATH))
        for etf, etf_data in all_data.items():
            update_excel(wb, etf, etf_data)
            print(f"  {etf} 已更新 ({len(etf_data)} 行)")
        wb.save(str(XLSX_PATH))
        print(f"已儲存至 {XLSX_PATH}")

    # 3. 印訊號
    print_signal(data)

    print("\n完成！")


if __name__ == "__main__":
    main()
